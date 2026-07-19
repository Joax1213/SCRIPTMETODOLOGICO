import os
import re
import logging
import base64
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .themes import get_theme, BASE_COLUMNS, QUALITY_COLUMNS

logger = logging.getLogger("bibliometric_analyzer")

def _truncate_at_word(text, max_chars):
    """Trunca el texto en el último límite de palabra antes de max_chars.
    Evita el corte de palabras a mitad que produce .{30,120} en regex.
    """
    if len(text) <= max_chars:
        return text
    truncated = text[:max_chars]
    last_space = truncated.rfind(' ')
    if last_space > max_chars // 2:  # asegurarse de no cortar demasiado pronto
        truncated = truncated[:last_space]
    return truncated.rstrip('.,;') + '...'

def render_mermaid_to_png(mermaid_code, output_png_path):
    """Renderiza un diagrama Mermaid a un archivo PNG utilizando la API pública de mermaid.ink."""
    try:
        import requests  # lazy import: solo se usa en esta función auxiliar
        mermaid_bytes = mermaid_code.encode("utf-8")
        base64_str = base64.urlsafe_b64encode(mermaid_bytes).decode("utf-8").rstrip("=")
        url = f"https://mermaid.ink/img/{base64_str}"
        response = requests.get(url, timeout=15)
        if response.status_code == 200:
            with open(output_png_path, "wb") as f:
                f.write(response.content)
            logger.info(f"Diagrama Mermaid renderizado y guardado exitosamente en: {output_png_path}")
            return True
        else:
            logger.warning(f"No se pudo renderizar Mermaid a PNG (Status de API: {response.status_code})")
    except Exception as e:
        logger.warning(f"Error al renderizar Mermaid a PNG vía API: {e}")
    return False


def generate_audit_matrix_template(output_path, theme="general"):
    """Genera una plantilla Excel de auditoría unificada según el tema seleccionado con estilos premium."""
    if not output_path:
        output_path = "plantilla_auditoria_sistematica.xlsx"
        
    logger.info(f"Generando plantilla Excel premium [Tema: {theme}] en: {output_path}")
    
    theme_spec = get_theme(theme)
    theme_cols = theme_spec["matrix_columns"]
    cols = BASE_COLUMNS + theme_cols + QUALITY_COLUMNS
    
    example_row = {col: "Dato del Paper" for col in cols}
    example_row["#"] = 1
    example_row["ID_OpenAlex"] = "https://openalex.org/W12345"
    example_row["DOI"] = "https://doi.org/10.xxxx/xxxxx"
    example_row["Título"] = "Título del artículo científico de ejemplo"
    example_row["Autores"] = "Ejemplo-Autor (2025)"
    example_row["Año"] = 2025
    example_row["Revista"] = "Nature Computational Science"
    example_row["Abstract"] = "Resumen del estudio..."
    example_row["Descubrimientos Principales"] = "Hallazgo cualitativo de ejemplo."
    example_row["Aporte al Tema"] = "Respaldo teórico para la justificación del estudio."
    
    theme_examples = theme_spec.get("example_row", {})
    for k, v in theme_examples.items():
        if k in cols:
            example_row[k] = v
            
    example_row["Calidad del Estudio (Alta/Media/Baja)"] = "Alta"
    example_row["Riesgo de Sesgo"] = "Bajo"
    example_row["Nivel de Evidencia"] = "Fuerte"

    df = pd.DataFrame([example_row], columns=cols)
    _write_premium_excel({"Auditoría Detallada": df}, output_path, theme)


def run_interactive_prisma_flow(output_path=None):
    print("==================================================")
    print("    ASISTENTE INTERACTIVO DE FLUJO PRISMA         ")
    print("==================================================")
    print("Por favor, introduce los conteos de registros solicitados:")
    
    try:
        # 1. db_records
        while True:
            try:
                db_records = int(input("1. Registros identificados en Bases de Datos (Scopus, PubMed, etc.): ") or 0)
                if db_records < 0:
                    print("   [Error] La cantidad de registros no puede ser negativa. Inténtalo de nuevo.")
                    continue
                break
            except ValueError:
                print("   [Error] Entrada no válida. Introduce un número entero.")

        # 2. other_records
        while True:
            try:
                other_records = int(input("2. Registros identificados en otras fuentes (Manual, etc.): ") or 0)
                if other_records < 0:
                    print("   [Error] La cantidad de registros no puede ser negativa. Inténtalo de nuevo.")
                    continue
                break
            except ValueError:
                print("   [Error] Entrada no válida. Introduce un número entero.")
        
        total_identified = db_records + other_records
        print(f"   -> Total registros identificados: {total_identified}")
        
        # 3. duplicates
        while True:
            try:
                duplicates = int(input("3. Registros duplicados eliminados: ") or 0)
                if duplicates < 0:
                    print("   [Error] La cantidad de duplicados no puede ser negativa. Inténtalo de nuevo.")
                    continue
                if duplicates > total_identified:
                    print(f"   [Error] Los duplicados ({duplicates}) no pueden exceder los registros identificados ({total_identified}). Inténtalo de nuevo.")
                    continue
                break
            except ValueError:
                print("   [Error] Entrada no válida. Introduce un número entero.")
                
        screened = total_identified - duplicates
        print(f"   -> Registros cribados (título/abstract): {screened}")
        
        # 4. screened_excluded
        while True:
            try:
                screened_excluded = int(input("4. Registros excluidos tras cribado de título/abstract: ") or 0)
                if screened_excluded < 0:
                    print("   [Error] La cantidad de excluidos no puede ser negativa. Inténtalo de nuevo.")
                    continue
                if screened_excluded > screened:
                    print(f"   [Error] Los excluidos en cribado ({screened_excluded}) no pueden exceder los registros cribados ({screened}). Inténtalo de nuevo.")
                    continue
                break
            except ValueError:
                print("   [Error] Entrada no válida. Introduce un número entero.")
                
        eligibility = screened - screened_excluded
        print(f"   -> Artículos evaluados en texto completo: {eligibility}")
        
        # 5. eligibility_excluded
        while True:
            try:
                eligibility_excluded = int(input("5. Artículos de texto completo EXCLUIDOS: ") or 0)
                if eligibility_excluded < 0:
                    print("   [Error] La cantidad de excluidos no puede ser negativa. Inténtalo de nuevo.")
                    continue
                if eligibility_excluded > eligibility:
                    print(f"   [Error] Los artículos excluidos ({eligibility_excluded}) no pueden exceder los artículos evaluados ({eligibility}). Inténtalo de nuevo.")
                    continue
                break
            except ValueError:
                print("   [Error] Entrada no válida. Introduce un número entero.")
        
        reasons = []
        if eligibility_excluded > 0:
            print("   Introduce las razones de exclusión del texto completo (deja vacío para terminar):")
            while True:
                reason = input("     - Razón de exclusión: ")
                if not reason:
                    break
                reasons.append(reason)
                
        included = eligibility - eligibility_excluded
        print(f"   -> Artículos incluidos finales: {included}")
        
        reasons_str = "; ".join(reasons) if reasons else "No cumple criterios de inclusión"
        
        mermaid_code = f"""flowchart TD
    classDef default fill:#f9f9f9,stroke:#333,stroke-width:1px;
    classDef highlight fill:#d8f3dc,stroke:#1b4332,stroke-width:2px;
    
    A["Registros identificados en bases de datos<br>N = {db_records}"] --> C
    B["Registros identificados en otras fuentes<br>N = {other_records}"] --> C
    C["Registros totales identificados<br>N = {total_identified}"] --> D
    
    D["Duplicados eliminados<br>N = {duplicates}"] --> E
    E["Registros únicos cribados<br>N = {screened}"] --> F
    
    F["Excluidos en cribado de título/abstract<br>N = {screened_excluded}"]
    E --> G["Artículos evaluados para elegibilidad (texto completo)<br>N = {eligibility}"]
    
    G --> H["Artículos de texto completo excluidos<br>N = {eligibility_excluded}<br>Razones: {reasons_str}"]
    G --> I["Artículos incluidos en la revisión sistemática<br>N = {included}"]:::highlight
"""
        
        print("\n==================================================")
        print("          DIAGRAMA MERMAID PRISMA GENERADO       ")
        print("==================================================")
        print(mermaid_code)
        
        if output_path:
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(mermaid_code)
            logger.info(f"Código del diagrama PRISMA guardado en: {output_path}")
            
            # Intentar compilar a PNG automáticamente en el mismo directorio
            base, _ = os.path.splitext(output_path)
            png_path = base + ".png"
            render_mermaid_to_png(mermaid_code, png_path)
            
    except Exception as e:
        logger.error(f"Error al ejecutar el flujo interactivo de PRISMA: {e}")


# — Placeholders que no cuentan como valores positivos de calidad —
_QUALITY_PLACEHOLDERS = frozenset({
    "no reportado", "revisar en texto completo", "por revisar",
    "n/a", "nan", "", "none", "not available",
})


def _is_positive(series, patterns):
    """True si el valor contiene alguno de los patrones con límite de palabra
    Y no es un placeholder que inflaría artificialmente el conteo de calidad.
    """
    combined = r'\b(?:' + '|'.join(re.escape(p) for p in patterns) + r')\b'
    not_placeholder = ~series.str.strip().str.lower().isin(_QUALITY_PLACEHOLDERS)
    return series.str.lower().str.contains(combined, regex=True, na=False) & not_placeholder


def parse_quality_and_bias(input_file):
    """Lee el Excel de auditoría y genera un bloque Markdown de calidad GRADE.
    Retorna (quality_markdown_str, quality_stats_dict) para que las RQs usen
    métricas reales del corpus.
    """
    bias_data = []
    if not input_file or not os.path.exists(input_file):
        return None, {}

    try:
        xl = pd.ExcelFile(input_file)
        sheet_name = None
        for name in xl.sheet_names:
            if "auditor" in name.lower() or "detallada" in name.lower():
                sheet_name = name
                break
        if not sheet_name:
            sheet_name = xl.sheet_names[0]

        df = pd.read_excel(input_file, sheet_name=sheet_name)

        # Buscar columnas de calidad exactas o altamente específicas para evitar colisión con columnas del tema
        col_sd   = [c for c in df.columns if "calidad del estudio" in c.lower()]
        if not col_sd:
            col_sd = [c for c in df.columns if "sd" in c.lower() or "calidad" in c.lower()]
            
        col_lod  = [c for c in df.columns if "riesgo de sesgo" in c.lower()]
        if not col_lod:
            col_lod = [c for c in df.columns if "lod" in c.lower() or "sesgo" in c.lower()]
            
        col_ctrl = [c for c in df.columns if c.lower() == "nivel de evidencia" or ("evidencia" in c.lower() and "grade" not in c.lower())]
        if not col_ctrl:
            col_ctrl = [c for c in df.columns if "control" in c.lower() or "réplic" in c.lower() or "evidencia" in c.lower()]

        if col_sd and col_lod and col_ctrl:
            c_sd, c_lod, c_ctrl = col_sd[0], col_lod[0], col_ctrl[0]

            # Patrones con límite de palabra: ya no hace match de la letra “s” suelta
            sd_patterns   = ["sí", "alta", "fuerte", "alto", "strong", "high", "yes"]
            lod_patterns  = ["sí", "bajo", "baja", "low", "none", "yes"]
            ctrl_patterns = ["sí", "alta", "fuerte", "moderada", "moderate", "high", "yes"]

            total_papers = len(df)
            sd_mask   = _is_positive(df[c_sd].astype(str),   sd_patterns)
            lod_mask  = _is_positive(df[c_lod].astype(str),  lod_patterns)
            ctrl_mask = _is_positive(df[c_ctrl].astype(str), ctrl_patterns)

            sd_si   = sd_mask.sum()
            lod_si  = lod_mask.sum()
            ctrl_si = ctrl_mask.sum()

            df['score'] = sd_mask.astype(int) + lod_mask.astype(int) + ctrl_mask.astype(int)

            score_avg  = df['score'].mean()
            strong_si  = (df['score'] == 3).sum()
            weak_si    = (df['score'] <= 1).sum()

            quality_summary = f"""
### Análisis de Calidad Científica y Sesgo (GRADE / Cochrane adaptada)

El análisis cualitativo de la muestra de **{total_papers} artículos** revela las siguientes métricas de consistencia científica:

*   **Validez de Reporte Estadístico / Calidad:** {sd_si} de {total_papers} artículos ({sd_si/total_papers*100:.1f}%) reportan análisis de calidad robustos o estadística detallada.
*   **Bajo Riesgo de Sesgo:** {lod_si} de {total_papers} artículos ({lod_si/total_papers*100:.1f}%) declaran calibración, controles de sesgo o límites claros.
*   **Consistencia de Controles / Evidencia:** {ctrl_si} de {total_papers} artículos ({ctrl_si/total_papers*100:.1f}%) muestran metodologías de control/replicabilidad adecuadas.

**Nivel de Calidad Promedio de la Muestra:** **{score_avg:.2f} / 3.00**
*   **Recomendación Fuerte (Bajo sesgo, Score = 3):** {strong_si} artículos ({strong_si/total_papers*100:.1f}%).
*   **Alto Riesgo de Sesgo (Score ≤ 1):** {weak_si} artículos ({weak_si/total_papers*100:.1f}%).
"""
            bias_data.append("| Autor (Áño) | Reporta Calidad | Bajo Sesgo | Controles / Evidencia | GRADE Score |")
            bias_data.append("|---|---|---|---|---|")

            col_author = [c for c in df.columns if "autor" in c.lower() or "año" in c.lower()]
            c_auth = col_author[0] if col_author else df.columns[1]

            for _, r in df.iterrows():
                auth     = r[c_auth]
                sd_val   = "Sí" if sd_mask.loc[r.name]   else "No"
                lod_val  = "Sí" if lod_mask.loc[r.name]  else "No"
                ctrl_val = "Sí" if ctrl_mask.loc[r.name] else "No"
                scr      = r['score']
                bias_data.append(f"| {auth} | {sd_val} | {lod_val} | {ctrl_val} | **{scr} / 3** |")

            quality_summary += "\n" + "\n".join(bias_data)

            # Devolver también las estadísticas para que generate_rqs_markdown las use
            quality_stats = {
                "total_papers" : total_papers,
                "sd_pct"       : sd_si   / total_papers * 100,
                "lod_pct"      : lod_si  / total_papers * 100,
                "ctrl_pct"     : ctrl_si / total_papers * 100,
                "score_avg"    : score_avg,
                "strong_si"    : int(strong_si),
                "strong_pct"   : strong_si / total_papers * 100,
                "weak_si"      : int(weak_si),
                "weak_pct"     : weak_si   / total_papers * 100,
            }
            return quality_summary, quality_stats

    except Exception as e:
        logger.error(f"Error al analizar calidad y sesgo en Excel: {e}")
    return None, {}

analyze_quality_bias_from_excel = parse_quality_and_bias

def generate_rqs_markdown(input_file, quality_stats=None):
    """Auto-genera preguntas de investigación (RQs) dinámicas a partir de las keywords más frecuentes del corpus.

    Args:
        input_file: Ruta al Excel de auditoría.
        quality_stats: Dict con métricas reales de parse_quality_and_bias (opcional).
                       Si se provee, RQ3 mostrará datos cuantitativos reales del corpus.
    """
    if not input_file or not input_file.lower().endswith(".xlsx"):
        return ""

    try:
        xl = pd.ExcelFile(input_file)
        sheet_name = None
        for name in xl.sheet_names:
            if "auditor" in name.lower() or "detallada" in name.lower():
                sheet_name = name
                break
        if not sheet_name:
            sheet_name = xl.sheet_names[0]

        df = pd.read_excel(input_file, sheet_name=sheet_name)

        text_content = ""
        for col_name in ["abstract", "titulo", "descubrimiento", "aporte", "tema", "Título", "Abstract"]:
            matching_cols = [c for c in df.columns if col_name in c.lower()]
            if matching_cols:
                text_content += " " + " ".join(df[matching_cols[0]].dropna().astype(str).tolist())
        text_content = text_content.lower()

        stopwords = {
            "the", "a", "an", "of", "in", "for", "and", "or", "to", "on", "at", "by", "with", "from",
            "is", "are", "was", "were", "this", "that", "these", "those", "it", "its", "be", "been",
            "study", "research", "analysis", "results", "effect", "effects", "using", "based", "used"
        }

        words = re.findall(r'\b[a-z]{4,}\b', text_content)
        freq = {}
        for w in words:
            if w not in stopwords:
                freq[w] = freq.get(w, 0) + 1

        sorted_kws = sorted(freq.items(), key=lambda x: -x[1])
        top_kws = [k.title() for k, v in sorted_kws[:6]]

        while len(top_kws) < 6:
            top_kws.append(f"TópicoClave{len(top_kws)+1}")

        k1, k2, k3, k4, k5, k6 = top_kws[:6]

        # Construir respuesta de RQ3 con datos reales si están disponibles
        if quality_stats and quality_stats.get("total_papers"):
            qs = quality_stats
            rq3_response = (
                f"El corpus incluye **{qs['strong_si']} artículos ({qs['strong_pct']:.1f}%)** con score GRADE = 3/3 "
                f"(bajo sesgo, controles adecuados y reporte estadístico robusto), "
                f"sobre un total de {qs['total_papers']} trabajos analizados. "
                f"El promedio de rigor metodológico es **{qs['score_avg']:.2f}/3.00**. "
                f"{qs['weak_si']} artículos ({qs['weak_pct']:.1f}%) presentan score ≤ 1 (alto riesgo de sesgo)."
            )
        else:
            rq3_response = "Consultar la Tabla de Calidad GRADE adjunta para evaluación detallada del corpus."

        rqs_md = f"""
### Preguntas de Investigación (Research Questions - RQs) Auto-generadas

El manuscrito y la revisión sistemática se estructuran formalmente en base a las siguientes preguntas metodológicas generadas automáticamente a partir de las palabras clave del corpus:

1.  **RQ1 (Estado del Arte):** ¿Cómo ha evolucionado la producción científica y la colaboración académica internacional sobre **{k1}** y su relación con **{k2}**?
    *   *Guía de análisis:* Examinar la distribución temporal de publicaciones, redes de co-autoría y fuentes principales que concentran investigación sobre **{k1}** en el periodo analizado.
2.  **RQ2 (Diseño Metodológico):** ¿Qué aproximaciones en el diseño experimental o tecnológico de **{k3}** han demostrado mayor consistencia para el análisis de **{k4}**?
    *   *Guía de análisis:* Comparar las columnas de diseño metodológico y técnicas en la matriz de síntesis. Identificar los métodos más recurrentes y su asociación con resultados positivos.
3.  **RQ3 (Rigor Científico):** ¿Cuál es el nivel de consistencia estadística y de reporte experimental en los estudios que investigan **{k5}**?
    *   *Respuesta basada en datos del corpus:* {rq3_response}
4.  **RQ4 (Limitaciones y Tendencias):** ¿Cuáles son las limitaciones críticas reportadas en torno a **{k6}** y qué perspectivas futuras se delinean para el sector?
    *   *Guía de análisis:* Revisar la columna de limitaciones de la matriz de síntesis y los aportes teóricos de los artículos de mayor calidad GRADE.
"""
        return rqs_md
    except Exception as e:
        logger.error(f"Error al generar RQs automáticas: {e}")
        return ""


def generate_populated_matrix(nodes, output_path, theme="general"):
    """Genera la matriz de síntesis de auditoría sistemática Excel poblada con datos y estilos premium."""
    if not output_path:
        output_path = "matriz_auditoria_sistematica.xlsx"
        
    logger.info(f"Generando matriz Excel premium poblada [Tema: {theme}] en: {output_path}")
    
    theme_spec = get_theme(theme)
    theme_cols = theme_spec["matrix_columns"]
    cols = BASE_COLUMNS + theme_cols + QUALITY_COLUMNS
    
    rows = []
    for idx, (node_id, data) in enumerate(nodes.items(), 1):
        title = data.get("Título", "")
        abstract = data.get("Abstract", "")
        texto_completo = data.get("TextoCompleto", "")
        search_text = title + " " + abstract + " " + (texto_completo or "")
        search_text_lower = search_text.lower()
        
        row = {col: "" for col in cols}
        row["#"] = idx
        row["ID_OpenAlex"] = data.get("ID", "")
        row["DOI"] = data.get("DOI", node_id)
        row["Título"] = title
        row["Autores"] = data.get("Autores", "Desconocido")
        row["Año"] = data.get("Año", "N/A")
        row["Revista"] = data.get("Revista", "N/A")
        row["Abstract"] = abstract
        row["Descubrimientos Principales"] = data.get("Descubrimientos Principales", "")
        row["Aporte al Tema"] = data.get("Aporte al Tema", "")
        
        used_sentences = set()
        for t_col in theme_cols:
            t_col_lower = t_col.lower()
            val = "No reportado"

            # ── Motor de extracción por tipo semántico de columna ──────────────────
            # 1. Concentración / Dosis
            if any(w in t_col_lower for w in ["concentración", "cantidad", "dosis", "valor", "concentracion", "dose"]):
                conc_pattern = r'(\d+(?:\.\d+)?(?:\s*(?:-|to|and|,\s*)\s*\d+(?:\.\d+)?)*\s*(?:mM|uM|µM|mg/L|g/L|ppm|mmol|%))'
                matches = re.findall(conc_pattern, search_text, re.IGNORECASE)
                if matches:
                    val = ", ".join(list(dict.fromkeys(matches))[:3])

            # 2. Rendimiento / Yield
            elif any(w in t_col_lower for w in ["rendimiento", "yield", "recuperación", "output"]):
                yield_pattern = r'(\d+(?:\.\d+)?(?:\s*(?:-|to|and|,\s*)\s*\d+(?:\.\d+)?)*\s*(?:mg/g|ug/g|g/kg|%|mg/100g))'
                matches = re.findall(yield_pattern, search_text, re.IGNORECASE)
                if matches:
                    val = ", ".join(list(dict.fromkeys(matches))[:3])

            # 3. Método / Técnica / Instrumento
            elif any(w in t_col_lower for w in ["método", "técnica", "instrumento", "metodología", "method", "technique", "instrument", "cuantificación", "extracción", "análisis"]):
                if any(w in search_text_lower for w in ["hplc", "chromatography", "cromatografía"]):
                    val = "HPLC-UV"
                elif any(w in search_text_lower for w in ["spectrometry", "espectrometría", "lc-ms"]):
                    val = "LC-MS/MS"
                elif any(w in search_text_lower for w in ["spectrophotometr", "espectrofotometría"]):
                    val = "Espectrofotometría"
                elif any(w in search_text_lower for w in ["rct", "randomized controlled", "ensayo controlado"]):
                    val = "Ensayo Controlado Aleatorizado (RCT)"
                elif any(w in search_text_lower for w in ["cohort", "cohorte"]):
                    val = "Estudio de Cohorte"
                elif any(w in search_text_lower for w in ["case-control", "caso-control", "caso control"]):
                    val = "Estudio Caso-Control"
                elif any(w in search_text_lower for w in ["cross-sectional", "transversal"]):
                    val = "Estudio Transversal"
                elif any(w in search_text_lower for w in ["systematic review", "meta-analysis", "revisión sistemática", "metaanálisis"]):
                    val = "Revisión Sistemática / Meta-análisis"
                elif any(w in search_text_lower for w in ["servqual", "holsat", "survey", "encuesta"]):
                    val = "Encuesta Estructurada (SERVQUAL)"
                elif any(w in search_text_lower for w in ["lean", "six sigma", "dmaic", "taguchi"]):
                    val = "Metodología Industrial (DMAIC / Taguchi)"
                elif any(w in search_text_lower for w in ["regression", "regresión", "logistic", "logística"]):
                    val = "Análisis de Regresión"
                else:
                    val = "Análisis Experimental"

            # 4. Especie / Variedad / Matriz / Tipo de estudio
            elif any(w in t_col_lower for w in ["especie", "variedad", "matriz", "alimento", "producto", "establecimiento", "destino", "species", "matrix", "tipo de estudio", "diseño", "diseno"]):
                if "vicia" in search_text_lower or "faba" in search_text_lower:
                    val = "Vicia faba L."
                elif "mucuna" in search_text_lower:
                    val = "Mucuna pruriens"
                elif any(w in search_text_lower for w in ["rct", "randomized controlled trial"]):
                    val = "Ensayo Controlado Aleatorizado (RCT)"
                elif any(w in search_text_lower for w in ["cohort study", "prospective cohort", "retrospective cohort"]):
                    val = "Estudio de Cohorte"
                elif any(w in search_text_lower for w in ["case-control", "caso control"]):
                    val = "Estudio Caso-Control"
                elif any(w in search_text_lower for w in ["cross-sectional", "transversal"]):
                    val = "Estudio Transversal"
                elif any(w in search_text_lower for w in ["systematic review", "meta-analysis"]):
                    val = "Revisión Sistemática"
                elif "hotel" in search_text_lower:
                    val = "Establecimiento Hotelero"
                elif "restaurant" in search_text_lower:
                    val = "Sector Restaurantes"
                elif "milk" in search_text_lower or "leche" in search_text_lower:
                    val = "Matriz Láctea"
                elif "wheat" in search_text_lower or "trigo" in search_text_lower:
                    val = "Matriz de Trigo"
                else:
                    val = "Modelo de Estudio"

            # 5. Resultados Estadísticos (OR / RR / HR / p-value) — columna específica de health_sciences
            elif any(w in t_col_lower for w in ["or/rr", "or/rr/hr", "p-value", "estadístico", "estadistico", "odds ratio", "hazard"]):
                stat_matches = re.findall(
                    r'(?:OR|RR|HR|odds ratio|risk ratio|hazard ratio)[\s=:]+([\d\.]+(?:\s*(?:\(|,|;)?\s*(?:IC|CI|95%)?[\s:]*[\d\.]+\s*[-–]\s*[\d\.]+\s*(?:\))?)?(?:,?\s*p\s*[<=]?\s*[\d\.]+)?)',
                    search_text, re.IGNORECASE)
                p_matches = re.findall(
                    r'p\s*[<=]\s*0\.\d+',
                    search_text, re.IGNORECASE)
                parts = []
                if stat_matches:
                    parts.append(stat_matches[0].strip())
                if p_matches:
                    parts.append(p_matches[0])
                val = "; ".join(parts[:2]) if parts else "No reportado"

            # 6. Intervención / Exposición — health_sciences
            elif any(w in t_col_lower for w in ["intervención", "intervencion", "exposición", "exposicion", "intervention", "exposure", "tratamiento"]):
                interv_patterns = [
                    r'(?:treated?\s+with|administered|received?|exposed?\s+to|intervención(?:\s+\w+){0,3})\s+([\w\-\s]+(?:\d+\s*(?:mg|g|mcg|IU|ml|ug))?)',
                    r'(?:mg|g|mcg|IU|dose)[^.]+',
                    r'(?:vs\.?|versus|compared\s+to|frente\s+a)[^.]+',
                ]
                target_search = abstract if (abstract and abstract.strip() not in ("", "nan", "Abstract no disponible.", "n/a")) else search_text
                for pat in interv_patterns:
                    m = re.search(pat, target_search, re.IGNORECASE)
                    if m:
                        snippet = m.group(0).strip()
                        if len(snippet) >= 10 and title.lower() not in snippet.lower():
                            val = _truncate_at_word(snippet, 90)
                            break

            # 7. Desenlace / Outcome
            elif any(w in t_col_lower for w in ["desenlace", "outcome", "endpoint", "result", "resultado"]):
                outcome_patterns = [
                    r'(?:primary\s+outcome|desenlace\s+primario|endpoint)[\s:]+([^.]+)',
                    r'(?:mortality|survival|incidence|prevalence|mortalidad|supervivencia|incidencia)[^.]+',
                    r'(?:reduce[ds]?|decrease[ds]?|improve[ds]?|reduj|mejor)[^.]+',
                ]
                for pat in outcome_patterns:
                    m = re.search(pat, search_text, re.IGNORECASE)
                    if m:
                        snippet = m.group(0).strip()
                        val = _truncate_at_word(snippet, 90)
                        break

            # 8. Población / Muestra
            elif any(w in t_col_lower for w in ["población", "poblacion", "muestra", "sujetos", "population", "sample", "paciente", "participant"]):
                pop_patterns = [
                    r'(\d+\s*(?:patients?|participants?|subjects?|adults?|children|pacientes?|sujetos?|individuos?)(?:[^.]+)?)',
                    r'(?:n\s*=\s*\d+[^.]*)',
                    r'(?:aged?\s+\d+[\s-]\d+|edad\s+(?:entre\s+)?\d+[^.]*)',
                ]
                for pat in pop_patterns:
                    m = re.search(pat, search_text, re.IGNORECASE)
                    if m:
                        snippet = m.group(0).strip()
                        val = _truncate_at_word(snippet, 90)
                        break

            # 9. Limitaciones (genérico, pero semántico)
            elif any(w in t_col_lower for w in ["limitación", "limitacion", "limitation", "limitante", "cuello", "debilidad", "sesgo", "bias", "bottleneck"]):
                limit_patterns = [
                    r'(?:limitation|limitación|debilidad|shortcoming|weakness|sin embargo|however)[^.]+',
                    r'(?:small\s+sample|muestra\s+pequeña|corto\s+seguimiento|short\s+follow)[^.]+',
                    r'(?:future\s+(?:studies?|research)|estudios?\ futuros?)[^.]+',
                ]
                for pat in limit_patterns:
                    m = re.search(pat, search_text, re.IGNORECASE)
                    if m:
                        snippet = m.group(0).strip()
                        val = _truncate_at_word(snippet, 100)
                        break
                else:
                    # Fallback fijo — solo si no hay texto de limitaciones real
                    fallbacks = [
                        "Seguimiento insuficiente / tamaño muestral limitado",
                        "Posible sesgo de selección o de información",
                        "Heterogeneidad de la muestra / variabilidad de resultados",
                        "Generalizabilidad limitada por contexto poblacional",
                    ]
                    val = fallbacks[idx % len(fallbacks)]

            # 9.5 Nivel de Evidencia (GRADE) — health_sciences / genérico
            elif any(w in t_col_lower for w in ["nivel de evidencia", "grade", "evidencia (grade)"]):
                grade_patterns = [
                    r'\b(alta|moderada|baja|muy\s+baja|high|moderate|low|very\s+low)\b',
                ]
                target_search = abstract if (abstract and abstract.strip() not in ("", "nan", "Abstract no disponible.", "n/a")) else search_text
                
                found_grade = None
                for pat in grade_patterns:
                    m = re.search(pat, target_search, re.IGNORECASE)
                    if m:
                        found_grade = m.group(0).strip().title()
                        # Normalizar a español
                        g_lower = found_grade.lower()
                        if "high" in g_lower: found_grade = "Alta"
                        elif "moderate" in g_lower: found_grade = "Moderada"
                        elif "very low" in g_lower: found_grade = "Muy baja"
                        elif "low" in g_lower: found_grade = "Baja"
                        break
                
                if found_grade:
                    val = found_grade
                else:
                    # Mapear por tipo de estudio a un nivel válido GRADE
                    design_patterns = [
                        (r'\b(systematic review|meta.analysis|rct|randomized controlled trial)\b', "Alta"),
                        (r'\b(cohort|case.control|prospective|retrospective)\b', "Moderada"),
                        (r'\b(cross.sectional|transversal|observational|survey)\b', "Baja"),
                    ]
                    for pat, grade_val in design_patterns:
                        if re.search(pat, target_search, re.IGNORECASE):
                            val = grade_val
                            break
                    else:
                        val = "Moderada"  # Fallback estándar


            # 10. Fallback genérico semántico: busca oración relevante por palabras clave de la columna
            else:
                col_kws = {
                    "variable": ["variable", "measure", "indicator", "metric", "kpi", "outcome", "dependent", "independent"],
                    "indicador": ["indicator", "kpi", "metric", "outcome", "measure"],
                    "hallazgo": ["result", "find", "show", "demonstrate", "observe", "conclude", "significant", "found"],
                    "ecosistema": ["ecosystem", "area", "region", "habitat", "forest", "river", "basin", "zone"],
                    "indicadores medidos": ["measured", "monitored", "assessed", "evaluated", "quantified"],
                    "implicación": ["implication", "policy", "recommend", "suggest", "propose"],
                    "proceso": ["process", "operation", "line", "stage", "step", "cycle"],
                    "kpi": ["kpi", "oee", "lead time", "defect", "throughput", "takt"],
                    "mejora": ["improvement", "reduction", "increase", "decrease", "optimize", "gain"],
                    "jurisdicción": ["country", "jurisdiction", "court", "law", "constitution", "statute"],
                    "país": ["country", "region", "nation", "economy", "GDP", "Latin America"],
                    "periodo": ["period", "year", "decade", "2010", "2015", "2020", "annual", "quarterly"],
                    "modelo": ["model", "regression", "GMM", "panel", "VAR", "OLS", "SEM"],
                }

                sentences = re.split(r'(?<=[.!?])\s+(?=[A-Z])', search_text)
                matched_kws = []
                for kw_key, kw_list in col_kws.items():
                    if kw_key in t_col_lower:
                        matched_kws = kw_list
                        break

                found_match = False
                if matched_kws:
                    for sent in sentences:
                        sent_clean = sent.strip()
                        if (any(w in sent_clean.lower() for w in matched_kws)
                                and len(sent_clean) > 20
                                and sent_clean not in used_sentences):
                            val = _truncate_at_word(sent_clean, 90)
                            used_sentences.add(sent_clean)
                            found_match = True
                            break

                if not found_match:
                    for sent in sentences:
                        sent_clean = sent.strip()
                        if len(sent_clean) > 20 and sent_clean not in used_sentences:
                            val = _truncate_at_word(sent_clean, 90)
                            used_sentences.add(sent_clean)
                            found_match = True
                            break

                if not found_match:
                    val = "Revisar en texto completo"

            row[t_col] = val
            
        # Inferir calidad, sesgo y evidencia a partir del contenido del abstract/texto
        _abs_lower = abstract.lower()
        _has_stats = any(w in _abs_lower for w in [
            "p < 0.05", "p<0.05", "significance", "confidence interval", "95% ci",
            "standard deviation", "mean ±", "anova", "regression", "odds ratio",
            "statistical", "p-value", "p value", "desviación estándar"
        ])
        _has_controls = any(w in _abs_lower for w in [
            "control group", "control negativo", "randomized", "blinded",
            "placebo", "replicate", "validated", "calibrated", "replicado", "calibrado"
        ])
        _low_bias = any(w in _abs_lower for w in [
            "randomized", "double-blind", "systematic review", "meta-analysis",
            "cochrane", "grade", "revisión sistemática", "ensayo clínico aleatorizado"
        ])
        row["Calidad del Estudio (Alta/Media/Baja)"] = "Alta" if (_has_stats and _has_controls) else ("Media" if _has_stats else "Por revisar")
        row["Riesgo de Sesgo"] = "Bajo" if _low_bias else ("Moderado" if _has_controls else "Por revisar")
        row["Nivel de Evidencia"] = "Fuerte" if _low_bias else ("Moderado" if _has_stats else "Limitado")
        
        rows.append(row)
        
    df_detail = pd.DataFrame(rows, columns=cols)
    
    rows_themes = []
    from .visualizer import auto_classify_axes, classify_node_by_keywords
    all_titles_abstracts = [data.get("Título", "") + " " + data.get("Abstract", "") for data in nodes.values()]
    all_kws = []
    for text in all_titles_abstracts:
        all_kws.extend(re.findall(r'\b[a-z]{4,}\b', text.lower()))
    axis_map, _ = auto_classify_axes(all_kws)
    
    for idx, (node_id, data) in enumerate(nodes.items(), 1):
        _, eje = classify_node_by_keywords(data.get("Título", ""), axis_map)
        rows_themes.append({
            "#": idx,
            "DOI": data.get("DOI", node_id),
            "Título": data.get("Título", ""),
            "Autores": data.get("Autores", ""),
            "Año": data.get("Año", ""),
            "Cluster Temático Asignado": eje
        })
    df_themes = pd.DataFrame(rows_themes)

    if not df_detail.empty:
        valid_years = pd.to_numeric(df_detail['Año'], errors='coerce').dropna().astype(int)
        rango_años = f"{valid_years.min()} - {valid_years.max()}" if not valid_years.empty else "N/A"
    else:
        rango_años = "N/A"
    _vc_revistas = df_detail['Revista'].value_counts() if not df_detail.empty else pd.Series(dtype=str)
    revista_top = _vc_revistas.index[0] if not _vc_revistas.empty else "N/A"
    
    resumen_data = {
        "Métrica Cienciométrica": [
            "Total de Artículos Auditados",
            "Rango Temporal de Publicación",
            "Revista / Fuente con Mayor Frecuencia",
            "Disciplina / Tema del Pipeline"
        ],
        "Valor Consolidado": [
            len(df_detail),
            rango_años,
            revista_top,
            theme_spec["name"]
        ]
    }
    df_resumen = pd.DataFrame(resumen_data)

    _write_premium_excel({
        "Auditoría Detallada": df_detail,
        "Resumen Ejecutivo": df_resumen,
        "Temas por Paper": df_themes
    }, output_path, theme)


def _write_premium_excel(sheet_dict, output_path, theme="general"):
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, df in sheet_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
        import openpyxl
        wb = openpyxl.load_workbook(output_path)
        
        header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        even_fill = PatternFill(start_color="F0F7F4", end_color="F0F7F4", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        for name in wb.sheetnames:
            ws = wb[name]
            ws.freeze_panes = "A2"
            
            max_col_letter = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{max_col_letter}{ws.max_row}"
            
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
            for row_idx in range(2, ws.max_row + 1):
                fill_to_use = even_fill if row_idx % 2 == 0 else white_fill
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = fill_to_use
                    cell.border = thin_border
                    cell.font = Font(name="Calibri", size=11)
                    if col_idx == 1:
                        cell.alignment = Alignment(horizontal="center")
                        
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)
                
        wb.save(output_path)
        logger.info(f"¡Estilos premium aplicados exitosamente a {output_path}!")
    except Exception as e:
        logger.error(f"Error al estilizar matriz Excel con openpyxl: {e}")
